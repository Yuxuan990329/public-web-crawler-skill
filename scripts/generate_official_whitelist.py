from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation


OUTPUT = "白名单网站模板_官方优化.xlsx"

HEADERS = [
    "site_name",
    "url",
    "domain",
    "category",
    "page_type",
    "enabled",
    "max_detail_pages",
    "note",
]

ROWS = [
    [
        "国家统计局-统计新闻",
        "https://www.stats.gov.cn/xw/tjxw/",
        "stats.gov.cn",
        "官方网站",
        "list",
        "yes",
        5,
        "候选栏目；用于替代首页，减少导航噪音",
    ],
    [
        "国家统计局-统计动态",
        "https://www.stats.gov.cn/xw/tjxw/tjdt/",
        "stats.gov.cn",
        "官方网站",
        "list",
        "yes",
        5,
        "候选栏目；新闻动态",
    ],
    [
        "国家统计局-通知公告",
        "https://www.stats.gov.cn/xw/tjxw/tzgg/",
        "stats.gov.cn",
        "官方网站",
        "list",
        "yes",
        5,
        "候选栏目；通知公告",
    ],
    [
        "国家统计局-数据解读",
        "https://www.stats.gov.cn/sj/zxfbhjd/",
        "stats.gov.cn",
        "官方网站",
        "list",
        "yes",
        5,
        "候选栏目；数据发布与解读",
    ],
    [
        "国家发改委",
        "https://www.ndrc.gov.cn/",
        "ndrc.gov.cn",
        "官方网站",
        "list",
        "no",
        5,
        "robots.txt 不允许当前爬虫访问；保留记录，不默认启用",
    ],
    [
        "中华人民共和国工业和信息化部",
        "https://www.miit.gov.cn",
        "miit.gov.cn",
        "官方网站",
        "list",
        "no",
        5,
        "robots.txt 不允许当前爬虫访问；保留记录，不默认启用",
    ],
]


def main():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "whitelist"
    sheet.append(HEADERS)
    for row in ROWS:
        sheet.append(row)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border

    widths = {
        "A": 34,
        "B": 48,
        "C": 20,
        "D": 18,
        "E": 12,
        "F": 12,
        "G": 18,
        "H": 50,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width

    validations = [
        (DataValidation(type="list", formula1='"官方网站,综合行业数据与研报平台,设计垂直平台"', allow_blank=False), "D2:D200"),
        (DataValidation(type="list", formula1='"page,list"', allow_blank=False), "E2:E200"),
        (DataValidation(type="list", formula1='"yes,no"', allow_blank=False), "F2:F200"),
        (DataValidation(type="whole", operator="between", formula1="0", formula2="50", allow_blank=False), "G2:G200"),
    ]
    for validation, cells in validations:
        sheet.add_data_validation(validation)
        validation.add(cells)

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:H{len(ROWS) + 1}"
    workbook.save(OUTPUT)
    print(OUTPUT)


if __name__ == "__main__":
    main()
