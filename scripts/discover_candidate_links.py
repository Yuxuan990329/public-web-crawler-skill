import argparse
import csv
import re
from dataclasses import dataclass
from datetime import datetime
from html.parser import HTMLParser
from pathlib import Path
from urllib.parse import urljoin, urlparse
from urllib import robotparser
from urllib.request import Request, urlopen

from openpyxl import load_workbook


DEFAULT_WHITELIST = "白名单网站模板.xlsx"
DEFAULT_OUTPUT_DIR = "outputs"
USER_AGENT = "CodexWhitelistCrawler/0.1 (+public whitelist research)"
REQUEST_TIMEOUT_SECONDS = 15

POSITIVE_KEYWORDS = [
    "政策",
    "新闻",
    "动态",
    "数据",
    "统计",
    "发布",
    "公告",
    "通知",
    "报告",
    "研究",
    "解读",
    "产业",
    "行业",
    "专题",
]

NEGATIVE_KEYWORDS = [
    "登录",
    "注册",
    "搜索",
    "邮箱",
    "订阅",
    "rss",
    "english",
    "新媒体",
    "信访",
    "智能问答",
    "无障碍",
    "地图",
    "关于我们",
]


@dataclass
class Site:
    site_name: str
    url: str
    domain: str
    category: str
    enabled: str


class LinkParser(HTMLParser):
    def __init__(self):
        super().__init__()
        self.links = []
        self._current_href = None
        self._current_text = []
        self._skip_depth = 0

    def handle_starttag(self, tag, attrs):
        tag = tag.lower()
        if tag in {"script", "style", "noscript", "svg"}:
            self._skip_depth += 1
            return
        if tag == "a":
            href = dict(attrs).get("href")
            if href:
                self._current_href = href
                self._current_text = []

    def handle_endtag(self, tag):
        tag = tag.lower()
        if tag in {"script", "style", "noscript", "svg"} and self._skip_depth:
            self._skip_depth -= 1
            return
        if tag == "a" and self._current_href:
            self.links.append((self._current_href, clean_text(" ".join(self._current_text))))
            self._current_href = None
            self._current_text = []

    def handle_data(self, data):
        if self._skip_depth or not self._current_href:
            return
        text = data.strip()
        if text:
            self._current_text.append(text)


def clean_text(value):
    return re.sub(r"\s+", " ", value or "").strip()


def normalize_domain(domain):
    domain = (domain or "").strip().lower()
    return domain[4:] if domain.startswith("www.") else domain


def url_domain(url):
    return normalize_domain(urlparse(url).netloc)


def is_same_host(url, base_url):
    return url_domain(url) == url_domain(base_url)


def read_sites(path):
    workbook = load_workbook(path, data_only=True)
    sheet = workbook["whitelist"] if "whitelist" in workbook.sheetnames else workbook.active
    headers = [str(cell.value).strip() if cell.value is not None else "" for cell in sheet[1]]
    indexes = {name: idx for idx, name in enumerate(headers)}
    sites = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row or not row[indexes["url"]]:
            continue
        sites.append(
            Site(
                site_name=str(row[indexes["site_name"]] or "").strip(),
                url=str(row[indexes["url"]] or "").strip(),
                domain=str(row[indexes["domain"]] or "").strip(),
                category=str(row[indexes["category"]] or "").strip(),
                enabled=str(row[indexes["enabled"]] or "no").strip().lower(),
            )
        )
    return sites


def split_site_names(value):
    return [name.strip() for name in re.split(r"[,，]", value or "") if name.strip()]


def filter_sites_by_names(sites, names):
    if not names:
        return sites
    wanted = set(names)
    return [site for site in sites if site.site_name in wanted]


def fetch_html(url):
    request = Request(url, headers={"User-Agent": USER_AGENT})
    with urlopen(request, timeout=REQUEST_TIMEOUT_SECONDS) as response:
        raw = response.read()
    for encoding in ("utf-8", "gb18030", "gbk"):
        try:
            return raw.decode(encoding)
        except UnicodeDecodeError:
            continue
    return raw.decode("utf-8", errors="ignore")


def can_fetch(url):
    parsed = urlparse(url)
    if parsed.scheme not in {"http", "https"}:
        return False
    robots_url = f"{parsed.scheme}://{parsed.netloc}/robots.txt"
    parser = robotparser.RobotFileParser()
    parser.set_url(robots_url)
    try:
        parser.read()
    except Exception:
        return True
    return parser.can_fetch(USER_AGENT, url)


def score_link(url, text):
    haystack = f"{text} {url}".lower()
    positives = [keyword for keyword in POSITIVE_KEYWORDS if keyword.lower() in haystack]
    negatives = [keyword for keyword in NEGATIVE_KEYWORDS if keyword.lower() in haystack]
    score = len(positives) * 10 - len(negatives) * 20
    if re.search(r"/20\d{2}/|20\d{2}", url):
        score += 3
    if len(text) >= 4:
        score += 1
    return score, positives, negatives


def discover(site, limit):
    if not can_fetch(site.url):
        raise RuntimeError("robots.txt 不允许访问")
    html = fetch_html(site.url)
    parser = LinkParser()
    parser.feed(html)

    seen = set()
    candidates = []
    for href, text in parser.links:
        absolute = urljoin(site.url, href)
        parsed = urlparse(absolute)
        if parsed.scheme not in {"http", "https"}:
            continue
        clean_url = parsed._replace(fragment="").geturl()
        if clean_url in seen:
            continue
        if not is_same_host(clean_url, site.url):
            continue
        if re.search(r"\.(pdf|doc|docx|xls|xlsx|zip|rar|jpg|jpeg|png|gif)(\?|$)", clean_url, re.I):
            continue
        score, positives, negatives = score_link(clean_url, text)
        if score <= 0:
            continue
        seen.add(clean_url)
        candidates.append(
            {
                "site_name": site.site_name,
                "category": site.category,
                "source_url": site.url,
                "candidate_url": clean_url,
                "link_text": text,
                "score": score,
                "positive_keywords": " ".join(positives),
                "negative_keywords": " ".join(negatives),
            }
        )
    candidates.sort(key=lambda row: row["score"], reverse=True)
    return candidates[:limit]


def output_path(output_dir):
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return Path(output_dir) / f"{timestamp}_候选栏目链接.csv"


def main():
    parser = argparse.ArgumentParser(description="从白名单首页发现更适合爬取的候选栏目链接。")
    parser.add_argument("--whitelist", default=DEFAULT_WHITELIST, help="白名单 Excel 路径。")
    parser.add_argument("--output-dir", default=DEFAULT_OUTPUT_DIR, help="CSV 输出目录。")
    parser.add_argument("--sites", default="", help="按 site_name 精确匹配多个站点，逗号分隔。")
    parser.add_argument("--limit-sites", type=int, default=0, help="只处理前 N 个启用站点，0 表示不限制。")
    parser.add_argument("--limit-links", type=int, default=20, help="每个站点最多输出 N 条候选链接。")
    args = parser.parse_args()

    sites = [site for site in read_sites(args.whitelist) if site.enabled == "yes"]
    sites = filter_sites_by_names(sites, split_site_names(args.sites))
    if args.limit_sites > 0:
        sites = sites[: args.limit_sites]

    rows = []
    for site in sites:
        try:
            rows.extend(discover(site, args.limit_links))
        except Exception as exc:
            rows.append(
                {
                    "site_name": site.site_name,
                    "category": site.category,
                    "source_url": site.url,
                    "candidate_url": "",
                    "link_text": "",
                    "score": 0,
                    "positive_keywords": "",
                    "negative_keywords": "",
                    "error": str(exc),
                }
            )

    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    path = output_path(output_dir)
    fields = [
        "site_name",
        "category",
        "source_url",
        "candidate_url",
        "link_text",
        "score",
        "positive_keywords",
        "negative_keywords",
        "error",
    ]
    with open(path, "w", newline="", encoding="utf-8-sig") as file:
        writer = csv.DictWriter(file, fieldnames=fields)
        writer.writeheader()
        for row in rows:
            writer.writerow(row)

    print(f"启用网站数: {len(sites)}")
    print(f"候选链接数: {len(rows)}")
    print(f"输出: {path}")


if __name__ == "__main__":
    main()
