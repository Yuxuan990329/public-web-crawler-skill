from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


OUTPUT = "白名单网站模板.xlsx"

HEADERS = [
    "site_name",
    "url",
    "domain",
    "category",
    "page_type",
    "enabled",
    "max_detail_pages",
    "note",
]

ROWS = [
    ["国家统计局", "https://www.stats.gov.cn/", "stats.gov.cn", "官方网站", "list", "yes", 5, "官方宏观数据源；可抓公开页面，但主题命中依赖具体栏目"],
    ["国家发改委", "https://www.ndrc.gov.cn/", "ndrc.gov.cn", "官方网站", "list", "no", 5, "robots.txt 不允许当前爬虫访问；保留来源记录，不进入自动抓取"],
    ["产业在线", "http://www.chinaioil.com/", "chinaioil.com", "官方网站", "list", "no", 5, "当前 DNS 解析失败；保留来源记录，待确认新域名或可访问栏目后再启用"],
    ["中华人民共和国工业和信息化部", "https://www.miit.gov.cn", "miit.gov.cn", "官方网站", "list", "no", 5, "robots.txt 不允许当前爬虫访问；保留来源记录，不进入自动抓取"],
    ["第一财经", "https://www.yicai.com", "yicai.com", "综合行业数据与研报平台", "list", "yes", 5, "公开资讯可抓；首页入口较宽，适合热点/宏观新闻，主题精确度依赖 matched-only 与 match-mode all"],
    ["智研咨询", "https://www.chyxx.com/", "chyxx.com", "综合行业数据与研报平台", "list", "yes", 5, "公开报告/百科可抓；首页入口较宽，需后续补行业栏目或搜索入口"],
    ["华经情报网", "https://www.huaon.com/", "huaon.com", "综合行业数据与研报平台", "list", "yes", 5, "公开报告/趋势页可抓；首页入口较宽，后续优先补家电/出海相关栏目"],
    ["艾瑞咨询", "https://www.iresearch.com.cn/", "iresearch.com.cn", "综合行业数据与研报平台", "list", "yes", 5, "完整报告需要付费；只抓公开页面"],
    ["发现报告", "https://www.fxbaogao.com/", "fxbaogao.com", "综合行业数据与研报平台", "list", "no", 5, "登录/收费限制；不在白名单保存账号密码，不做登录抓取；仅在提供公开 URL 时复测"],
    ["艾媒网", "https://www.iimedia.cn/", "iimedia.cn", "综合行业数据与研报平台", "list", "no", 5, "疑似登录/收费限制；基础模式不抓，后续仅处理公开免费页面"],
    ["前瞻网", "https://bg.qianzhan.com/report/chuhai/", "qianzhan.com", "综合行业数据与研报平台", "list", "yes", 5, "出海专题报告栏目；完整报告收费，只抓公开页面和报告简介"],
    ["CBNData", "https://www.cbndata.com/", "cbndata.com", "综合行业数据与研报平台", "list", "yes", 5, "公开资讯/报告页可抓；首页入口较宽，需后续补报告栏目或搜索入口"],
    ["36氪", "https://36kr.com", "36kr.com", "综合行业数据与研报平台", "list", "yes", 5, "首页未发现可抓详情链接；疑似前端动态渲染，需浏览器核对具体栏目，不绕过反爬"],
    ["洞见研报", "https://www.djyanbao.com/index", "djyanbao.com", "综合行业数据与研报平台", "list", "yes", 5, "首页未发现可抓详情链接；需补公开栏目 URL 或浏览器核对"],
    ["亿欧", "https://www.iyiou.com/", "iyiou.com", "综合行业数据与研报平台", "list", "yes", 5, "首页未发现可抓详情链接；需补公开栏目 URL 或浏览器核对"],
]


def style_sheet(ws, rows_count):
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border

    widths = {
        "A": 30,
        "B": 42,
        "C": 24,
        "D": 30,
        "E": 14,
        "F": 12,
        "G": 18,
        "H": 42,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{rows_count}"


def add_validations(ws):
    validations = [
        (DataValidation(type="list", formula1='"官方网站,综合行业数据与研报平台,设计垂直平台"', allow_blank=False), "D2:D200"),
        (DataValidation(type="list", formula1='"page,list"', allow_blank=False), "E2:E200"),
        (DataValidation(type="list", formula1='"yes,no"', allow_blank=False), "F2:F200"),
        (DataValidation(type="whole", operator="between", formula1="0", formula2="50", allow_blank=False), "G2:G200"),
    ]
    for validation, cells in validations:
        ws.add_data_validation(validation)
        validation.add(cells)


def add_readme(wb):
    ws = wb.create_sheet("README")
    rows = [
        ["字段", "说明"],
        ["site_name", "网站名称，供输出和日志识别使用。"],
        ["url", "允许访问的页面或入口页。Skill 只能访问白名单内 URL 及同域名详情页。"],
        ["domain", "允许访问的域名，用于安全校验。"],
        ["category", "网站分类，可从下拉选项选择。"],
        ["page_type", "page 表示普通页面；list 表示列表页，允许继续抓同域名详情页。"],
        ["enabled", "yes 表示启用；no 表示暂不访问。"],
        ["max_detail_pages", "列表页最多抓取详情页数量，第一版默认 5。"],
        ["note", "访问限制、收费说明、反爬说明等备注。"],
    ]
    for row in rows:
        ws.append(row)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 82
    ws.freeze_panes = "A2"


def main():
    wb = Workbook()
    ws = wb.active
    ws.title = "whitelist"
    ws.append(HEADERS)
    for row in ROWS:
        ws.append(row)

    style_sheet(ws, len(ROWS) + 1)
    add_validations(ws)
    add_readme(wb)
    wb.save(OUTPUT)
    print(OUTPUT)


if __name__ == "__main__":
    main()
