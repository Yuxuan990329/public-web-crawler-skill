import argparse
import csv
import http.client
import io
import json
import ipaddress
import re
import socket
import ssl
import time
import uuid
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Iterable
from urllib import robotparser
from urllib.error import HTTPError, URLError
from urllib.parse import parse_qsl, quote, urlencode, urljoin, urlparse, urlsplit, urlunsplit
from urllib.request import HTTPHandler, HTTPRedirectHandler, HTTPSHandler, ProxyHandler, Request, build_opener, urlopen

from openpyxl import load_workbook

from crawler_defaults import DEFAULT_WHITELIST
from extractors import clean_text, extract_page
from summarize import PROVIDERS, fallback_summary, set_summary_config, set_summary_mode, summarize_enhancement


DEFAULT_OUTPUT_DIR = "outputs"
DEFAULT_LOG_DIR = "logs"
USER_AGENT = "CodexWhitelistCrawler/0.1 (+public whitelist research)"
REQUEST_TIMEOUT_SECONDS = 15
REQUEST_DELAY_SECONDS = 1.5
KNOWN_TLS_CHAIN_LIMIT_DOMAINS = {"qianzhan.com"}
KNOWN_TLS_CHAIN_ERROR_CODES = {20, 21}

CLASSIFICATION_KEYWORDS = {
    "政策": ["政策", "通知", "公告", "规划", "条例", "办法", "意见"],
    "报告": ["报告", "白皮书", "研究", "研判", "调研"],
    "数据": ["数据", "统计", "指数", "规模", "同比", "环比"],
    "新闻": ["新闻", "动态", "发布", "召开", "调研"],
    "产业": ["产业", "行业", "市场", "企业", "链"],
}


@dataclass
class Site:
    site_name: str
    url: str
    domain: str
    category: str
    page_type: str
    enabled: str
    max_detail_pages: int
    note: str


def normalize_domain(domain):
    domain = (domain or "").strip().lower()
    return domain[4:] if domain.startswith("www.") else domain


def url_domain(url):
    return normalize_domain(urlparse(url).netloc)


def is_same_or_subdomain(url, domain):
    target = url_domain(url)
    allowed = normalize_domain(domain)
    return target == allowed or target.endswith("." + allowed)


def request_safe_url(url):
    parts = urlsplit(url)
    return urlunsplit(
        (
            parts.scheme,
            parts.netloc,
            quote(parts.path, safe="/%"),
            quote(parts.query, safe="=&?/:+,%"),
            parts.fragment,
        )
    )


def read_sites(path):
    workbook = load_workbook(path, data_only=True)
    sheet = workbook["whitelist"] if "whitelist" in workbook.sheetnames else workbook.active
    headers = [str(cell.value).strip() if cell.value is not None else "" for cell in sheet[1]]
    indexes = {name: idx for idx, name in enumerate(headers)}
    required = ["site_name", "url", "domain", "category", "page_type", "enabled", "max_detail_pages", "note"]
    missing = [name for name in required if name not in indexes]
    if missing:
        raise ValueError(f"白名单缺少字段: {', '.join(missing)}")

    sites = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row or not row[indexes["url"]]:
            continue
        max_detail_pages = row[indexes["max_detail_pages"]]
        try:
            max_detail_pages = int(max_detail_pages)
        except (TypeError, ValueError):
            max_detail_pages = 5
        sites.append(
            Site(
                site_name=str(row[indexes["site_name"]] or "").strip(),
                url=str(row[indexes["url"]] or "").strip(),
                domain=str(row[indexes["domain"]] or "").strip(),
                category=str(row[indexes["category"]] or "").strip(),
                page_type=str(row[indexes["page_type"]] or "page").strip().lower(),
                enabled=str(row[indexes["enabled"]] or "no").strip().lower(),
                max_detail_pages=max_detail_pages,
                note=str(row[indexes["note"]] or "").strip(),
            )
        )
    return sites


def split_site_names(value):
    return [name.strip() for name in re.split(r"[,，]", value or "") if name.strip()]


def filter_sites_by_names(sites, names):
    if not names:
        return sites
    wanted = set(names)
    return [site for site in sites if site.site_name in wanted]


def split_keywords(topic):
    return [part.strip() for part in re.split(r"\s+", topic or "") if part.strip()]


def matched_keywords(title, content, keywords):
    haystack = f"{title} {content}".lower()
    return [keyword for keyword in keywords if keyword.lower() in haystack]


def classify_content(title, content):
    haystack = f"{title} {content[:1000]}"
    scores = {}
    for category, keywords in CLASSIFICATION_KEYWORDS.items():
        scores[category] = sum(1 for keyword in keywords if keyword in haystack)
    best_category, best_score = max(scores.items(), key=lambda item: item[1])
    return best_category if best_score > 0 else "其他"


def extract_publish_date(title, content, url):
    haystack = f"{url} {title} {content[:1000]}"
    patterns = [
        r"(20\d{2})[-/.年](\d{1,2})[-/.月](\d{1,2})",
        r"(20\d{2})(\d{2})(\d{2})",
    ]
    for pattern in patterns:
        for match in re.finditer(pattern, haystack):
            try:
                year, month, day = (int(part) for part in match.groups())
                return date(year, month, day).isoformat()
            except ValueError:
                continue
    return ""


def parse_date(value, field_name):
    if not value:
        return None
    try:
        return date.fromisoformat(value)
    except ValueError as exc:
        raise ValueError(f"{field_name} 日期格式应为 YYYY-MM-DD") from exc


def in_date_range(publish_date, date_from=None, date_to=None):
    if not publish_date or (date_from is None and date_to is None):
        return True
    current = date.fromisoformat(publish_date)
    if date_from and current < date_from:
        return False
    if date_to and current > date_to:
        return False
    return True


def looks_blocked(title, content):
    text = f"{title} {content}".lower()
    markers = ["验证码", "请登录", "請登入", "登录后", "扫码登录", "403 forbidden", "access denied", "captcha"]
    return any(marker.lower() in text for marker in markers)


def can_fetch(url, user_agent):
    parsed = urlparse(url)
    if parsed.scheme not in {"http", "https"}:
        return False
    robots_url = f"{parsed.scheme}://{parsed.netloc}/robots.txt"
    parser = robotparser.RobotFileParser()
    parser.set_url(robots_url)
    try:
        with safe_open(robots_url, parsed.hostname) as response:
            raw = response.read()
        parser.parse(raw.decode("utf-8", errors="replace").splitlines())
    except UnsafeTargetError:
        return False
    except HTTPError as exc:
        exc.close()
        return exc.code in {404, 410}
    except URLError as exc:
        if tls_verification_failure(exc):
            raise
        return False
    except (TimeoutError, OSError):
        return False
    return parser.can_fetch(user_agent, url)


class UnsafeTargetError(ValueError):
    pass


def _is_public_ip(value):
    ip = ipaddress.ip_address(value)
    if not ip.is_global or ip.is_multicast or getattr(ip, "is_site_local", False):
        return False
    embedded = [getattr(ip, "ipv4_mapped", None), getattr(ip, "sixtofour", None)]
    teredo = getattr(ip, "teredo", None)
    if teredo:
        embedded.extend(teredo)
    for candidate in embedded:
        if candidate is not None and not _is_public_ip(str(candidate)):
            return False
    for prefix in (ipaddress.ip_network("64:ff9b::/96"), ipaddress.ip_network("64:ff9b:1::/48")):
        if ip.version == 6 and ip in prefix:
            mapped = ipaddress.ip_address(int(ip) & 0xFFFFFFFF)
            if not _is_public_ip(str(mapped)):
                return False
    return True


def resolve_public_addresses(host, port):
    try:
        addresses = socket.getaddrinfo(host, port, type=socket.SOCK_STREAM)
    except socket.gaierror as exc:
        raise UnsafeTargetError(f"无法解析目标域名: {host}") from exc
    if not addresses or any(not _is_public_ip(item[4][0]) for item in addresses):
        raise UnsafeTargetError("目标域名解析到非公网 IP，已拒绝请求")
    return addresses


def validate_public_target(url, allowed_domain):
    parsed = urlparse(url)
    if parsed.scheme not in {"http", "https"}:
        raise UnsafeTargetError("仅允许 HTTP/HTTPS 请求")
    if not is_same_or_subdomain(url, allowed_domain):
        raise UnsafeTargetError("重定向目标不匹配白名单域名")
    host = parsed.hostname
    if not host:
        raise UnsafeTargetError("请求 URL 缺少主机名")
    resolve_public_addresses(host, parsed.port or (443 if parsed.scheme == "https" else 80))


def _create_public_connection(host, port, timeout, source_address=None):
    addresses = resolve_public_addresses(host, port)
    last_error = None
    for family, socktype, proto, _canonname, sockaddr in addresses:
        sock = None
        try:
            sock = socket.socket(family, socktype, proto)
            sock.settimeout(timeout)
            if source_address:
                sock.bind(source_address)
            sock.connect(sockaddr)
            return sock
        except OSError as exc:
            last_error = exc
            if sock is not None:
                sock.close()
    if last_error is not None:
        raise last_error
    raise UnsafeTargetError("目标域名没有可用的公网地址")


class PublicHTTPConnection(http.client.HTTPConnection):
    def connect(self):
        self.sock = _create_public_connection(self.host, self.port, self.timeout, self.source_address)
        if self._tunnel_host:
            raise UnsafeTargetError("禁止通过代理隧道访问白名单站点")


class PublicHTTPSConnection(http.client.HTTPSConnection):
    def connect(self):
        self.sock = _create_public_connection(self.host, self.port, self.timeout, self.source_address)
        if self._tunnel_host:
            raise UnsafeTargetError("禁止通过代理隧道访问白名单站点")
        self.sock = self._context.wrap_socket(self.sock, server_hostname=self.host)


class PublicHTTPHandler(HTTPHandler):
    def http_open(self, req):
        return self.do_open(PublicHTTPConnection, req)


class PublicHTTPSHandler(HTTPSHandler):
    def https_open(self, req):
        return self.do_open(PublicHTTPSConnection, req, context=self._context)


class SafeRedirectHandler(HTTPRedirectHandler):
    def __init__(self, allowed_domain):
        super().__init__()
        self.allowed_domain = allowed_domain

    def redirect_request(self, req, fp, code, msg, headers, newurl):
        target = urljoin(req.full_url, newurl)
        validate_public_target(target, self.allowed_domain)
        return super().redirect_request(req, fp, code, msg, headers, target)


def safe_open(url, allowed_domain):
    validate_public_target(url, allowed_domain)
    request = Request(request_safe_url(url), headers={"User-Agent": USER_AGENT})
    opener = build_opener(
        ProxyHandler({}),
        PublicHTTPHandler(),
        PublicHTTPSHandler(),
        SafeRedirectHandler(allowed_domain),
    )
    response = opener.open(request, timeout=REQUEST_TIMEOUT_SECONDS)
    final_url = response.geturl()
    try:
        validate_public_target(final_url, allowed_domain)
    except Exception:
        response.close()
        raise
    return response


def fetch_html(url, allowed_domain=None):
    allowed_domain = allowed_domain or url_domain(url)
    with safe_open(url, allowed_domain) as response:
        content_type = response.headers.get("Content-Type", "")
        if "text/html" not in content_type and "application/xhtml+xml" not in content_type:
            raise ValueError(f"非 HTML 内容: {content_type}")
        raw = response.read()
    for encoding in ("utf-8", "gb18030", "gbk"):
        try:
            return raw.decode(encoding)
        except UnicodeDecodeError:
            continue
    return raw.decode("utf-8", errors="ignore")


def fetch_pdf_text(url, allowed_domain=None):
    allowed_domain = allowed_domain or url_domain(url)
    with safe_open(url, allowed_domain) as response:
        content_type = response.headers.get("Content-Type", "")
        if "pdf" not in content_type.lower() and not url.lower().endswith(".pdf"):
            raise ValueError(f"非 PDF 内容: {content_type}")
        raw = response.read()
    try:
        from pypdf import PdfReader
    except ImportError as exc:
        raise ValueError("PDF 正文提取依赖 pypdf 未安装") from exc
    reader = PdfReader(io.BytesIO(raw))
    parts = []
    for page in reader.pages:
        parts.append(page.extract_text() or "")
    return clean_text(" ".join(parts))


def same_domain_links(base_url, links, domain, limit, include_pdfs=False):
    seen = set()
    results = []
    base_host = url_domain(base_url)
    for link, link_text in links:
        absolute = urljoin(base_url, link)
        parsed = urlparse(absolute)
        if parsed.scheme not in {"http", "https"}:
            continue
        clean_url = parsed._replace(fragment="").geturl()
        if clean_url in seen:
            continue
        if not is_same_or_subdomain(clean_url, domain):
            continue
        if url_domain(clean_url) != base_host:
            continue
        if re.search(r"\.(doc|docx|xls|xlsx|zip|rar|jpg|jpeg|png|gif)(\?|$)", clean_url, re.I):
            continue
        if not include_pdfs and re.search(r"\.pdf(\?|$)", clean_url, re.I):
            continue
        if should_skip_detail_link(clean_url, link_text):
            continue
        seen.add(clean_url)
        results.append(clean_url)
        if len(results) >= limit:
            break
    return results


def should_skip_detail_link(url, link_text):
    text = (link_text or "").strip().lower()
    lowered_url = url.lower()
    if is_same_or_subdomain(url, "qianzhan.com") and "/report/detail/" not in lowered_url:
        return True
    blocked_url_parts = [
        "/znwd/",
        "/wzgl/",
        "/english/",
        "/rss",
        "/hd/xmt/",
        "login",
        "register",
        "search",
    ]
    blocked_text_parts = [
        "rss",
        "english",
        "新媒体",
        "信访",
        "智能问答",
        "无障碍",
        "登录",
        "注册",
        "搜索",
        "邮箱",
        "订阅",
        "可行性研究报告",
        "专精特新申报",
        "市场地位证明",
        "商业计划书",
        "定制报告",
        "提交需求",
    ]
    if any(part in lowered_url for part in blocked_url_parts):
        return True
    if any(part.lower() in text for part in blocked_text_parts):
        return True
    if len(text) < 6 and not re.search(r"/20\d{2}/|20\d{2}", lowered_url):
        return True
    return False


def result_row(
    site,
    topic,
    url,
    source_type,
    title="",
    content="",
    status="ok",
    error="",
    keywords=None,
    extractor="",
    publish_date="",
    content_type="html",
    candidate_title="",
    candidate_snippet="",
    score="",
    source_stage="",
    candidate_reason="",
    known_limit="",
    quality_issue="",
    review_required="",
    ai_relevance_score="",
    ai_keep="",
    ai_filter_reason="",
):
    keywords = keywords or []
    summary = ""
    ai_summary = ""
    ai_category = ""
    ai_reason = ""
    if status in {"ok", "matched"} and keywords:
        summary = fallback_summary(" ".join((content or "").split()))
        ai_result = summarize_enhancement(title=title, url=url, content=content, topic=topic, keywords=keywords)
        ai_summary = ai_result.get("ai_summary", "")
        ai_category = ai_result.get("ai_category", "")
        ai_reason = ai_result.get("ai_reason", "")
    return {
        "title": title,
        "url": url,
        "site_name": site.site_name,
        "crawl_time": datetime.now().isoformat(timespec="seconds"),
        "topic": topic,
        "content": content,
        "summary": summary,
        "ai_summary": ai_summary,
        "ai_category": ai_category,
        "ai_reason": ai_reason,
        "category": classify_content(title, content) if content else "",
        "publish_date": publish_date,
        "status": status,
        "error": error,
        "matched_keywords": " ".join(keywords),
        "source_type": source_type,
        "content_type": content_type,
        "extractor": extractor,
        "candidate_title": candidate_title,
        "candidate_snippet": candidate_snippet,
        "score": score,
        "source_stage": source_stage,
        "candidate_reason": candidate_reason,
        "known_limit": known_limit,
        "quality_issue": quality_issue,
        "review_required": review_required,
        "ai_relevance_score": ai_relevance_score,
        "ai_keep": ai_keep,
        "ai_filter_reason": ai_filter_reason,
    }


def tls_verification_failure(exc):
    reason = exc.reason if isinstance(exc, URLError) else exc
    if not isinstance(reason, ssl.SSLCertVerificationError):
        return None
    code = getattr(reason, "verify_code", None)
    message = clean_text(getattr(reason, "verify_message", "") or str(reason))
    return code, message


def is_known_tls_chain_limit(url, code, message):
    domain = url_domain(url)
    known_domain = any(domain == allowed or domain.endswith("." + allowed) for allowed in KNOWN_TLS_CHAIN_LIMIT_DOMAINS)
    return known_domain and code in KNOWN_TLS_CHAIN_ERROR_CODES


def tls_failure_row(site, topic, url, source_type, exc):
    tls_failure = tls_verification_failure(exc)
    if not tls_failure:
        return None
    code, message = tls_failure
    if is_known_tls_chain_limit(url, code, message):
        return result_row(
            site,
            topic,
            url,
            source_type,
            status="skipped",
            error=f"tls_certificate_chain_unavailable: {code or 'unknown'} {message}",
            known_limit="TLS证书链不可用，已保持证书校验并跳过抓取",
        )
    return result_row(
        site,
        topic,
        url,
        source_type,
        status="failed",
        error=f"tls_certificate_rejected: {code or 'unknown'} {message}",
        quality_issue="TLS证书校验失败，需核验站点或本机信任链",
        review_required="yes",
    )


def crawl_url(site, topic, keywords, url, source_type, match_mode="any", date_from=None, date_to=None, include_pdfs=False):
    if not is_same_or_subdomain(url, site.domain):
        return result_row(site, topic, url, source_type, status="skipped", error="URL 不匹配白名单域名")
    try:
        robots_allowed = can_fetch(url, USER_AGENT)
    except URLError as exc:
        tls_row = tls_failure_row(site, topic, url, source_type, exc)
        if tls_row:
            return tls_row
        return result_row(site, topic, url, source_type, status="skipped", error="robots.txt 获取失败，已拒绝访问")
    if not robots_allowed:
        return result_row(site, topic, url, source_type, status="skipped", error="robots.txt 不允许访问")
    try:
        if re.search(r"\.pdf(\?|$)", url, re.I):
            if not include_pdfs:
                return result_row(site, topic, url, source_type, status="skipped", error="PDF 提取未开启", content_type="pdf")
            title = clean_text(url.rsplit("/", 1)[-1])
            content = fetch_pdf_text(url, site.domain)
            links = []
            extractor = "pdf"
            content_type = "pdf"
        else:
            html = fetch_html(url, site.domain)
            page = extract_page(url, html)
            title, content, links = page.title, page.content, page.links
            extractor = page.extractor
            content_type = "html"
    except HTTPError as exc:
        return result_row(site, topic, url, source_type, status="failed", error=f"HTTP {exc.code}")
    except URLError as exc:
        tls_row = tls_failure_row(site, topic, url, source_type, exc)
        if tls_row:
            return tls_row
        return result_row(site, topic, url, source_type, status="failed", error=str(exc))
    except (TimeoutError, ValueError) as exc:
        return result_row(site, topic, url, source_type, status="failed", error=str(exc))

    if not clean_text(content):
        return result_row(
            site,
            topic,
            url,
            source_type,
            title=title,
            status="failed",
            error="正文为空",
            extractor=extractor,
            content_type=content_type,
            quality_issue="正文为空",
            review_required="yes",
        )

    publish_date = extract_publish_date(title, content, url)
    if publish_date and not in_date_range(publish_date, date_from, date_to):
        return result_row(
            site,
            topic,
            url,
            source_type,
            title=title,
            content=content[:500],
            status="skipped",
            error="发布时间不在范围内",
            extractor=extractor,
            publish_date=publish_date,
            content_type=content_type,
        )

    if looks_blocked(title, content):
        return result_row(
            site,
            topic,
            url,
            source_type,
            title=title,
            content=content[:500],
            status="failed",
            error="疑似登录、验证码或访问限制",
            extractor=extractor,
            publish_date=publish_date,
            content_type=content_type,
        )

    matched = matched_keywords(title, content, keywords)
    # any：任一关键词命中即算匹配；all：所有关键词都必须命中
    if match_mode == "all":
        is_matched = len(matched) == len(keywords) and bool(keywords)
    else:
        is_matched = bool(matched)
    status = "matched" if is_matched else "unmatched"
    return result_row(
        site,
        topic,
        url,
        source_type,
        title=title,
        content=content,
        status=status,
        keywords=matched,
        extractor=extractor,
        publish_date=publish_date,
        content_type=content_type,
    ), links


def iter_crawl_results(
    sites,
    topic,
    keywords,
    include_list_pages=False,
    match_mode="any",
    date_from=None,
    date_to=None,
    include_pdfs=False,
):
    visited_urls = set()
    for site in sites:
        if site.enabled != "yes":
            continue
        if not site.domain or not site.url:
            yield result_row(site, topic, site.url, "page", status="failed", error="白名单缺少 domain 或 url")
            continue

        if site.url in visited_urls:
            yield result_row(site, topic, site.url, "page", status="skipped", error="URL 已抓取，跳过去重")
            continue
        visited_urls.add(site.url)
        page_result = crawl_url(
            site,
            topic,
            keywords,
            site.url,
            "page",
            match_mode=match_mode,
            date_from=date_from,
            date_to=date_to,
            include_pdfs=include_pdfs,
        )
        time.sleep(REQUEST_DELAY_SECONDS)

        if isinstance(page_result, tuple):
            row, links = page_result
            if site.page_type != "list" or include_list_pages:
                yield row
        else:
            yield page_result
            continue

        if site.page_type != "list" or site.max_detail_pages <= 0:
            continue

        detail_links = same_domain_links(site.url, links, site.domain, site.max_detail_pages, include_pdfs=include_pdfs)
        if not detail_links and not include_list_pages:
            yield result_row(site, topic, site.url, "page", status="skipped", error="列表页未发现可抓取详情链接")
            continue

        emitted_detail = False
        for detail_url in detail_links:
            if detail_url in visited_urls:
                continue
            visited_urls.add(detail_url)
            detail_result = crawl_url(
                site,
                topic,
                keywords,
                detail_url,
                "detail",
                match_mode=match_mode,
                date_from=date_from,
                date_to=date_to,
                include_pdfs=include_pdfs,
            )
            time.sleep(REQUEST_DELAY_SECONDS)
            if isinstance(detail_result, tuple):
                yield detail_result[0]
            else:
                yield detail_result
            emitted_detail = True

        if not emitted_detail and not include_list_pages:
            yield result_row(site, topic, site.url, "page", status="skipped", error="详情链接均已抓取，跳过去重")


def write_csv(rows: Iterable[dict], output_path):
    fields = [
        "title",
        "url",
        "site_name",
        "crawl_time",
        "topic",
        "content",
        "summary",
        "ai_summary",
        "ai_category",
        "ai_reason",
        "category",
        "publish_date",
        "status",
        "error",
        "matched_keywords",
        "source_type",
        "content_type",
        "extractor",
        "candidate_title",
        "candidate_snippet",
        "score",
        "source_stage",
        "candidate_reason",
        "known_limit",
        "quality_issue",
        "review_required",
        "ai_relevance_score",
        "ai_keep",
        "ai_filter_reason",
    ]
    with open(output_path, "w", newline="", encoding="utf-8-sig") as file:
        writer = csv.DictWriter(file, fieldnames=fields)
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def filter_matched(rows: Iterable[dict]):
    for row in rows:
        if row.get("status") == "matched":
            yield row


def canonical_url(url):
    parsed = urlparse(url or "")
    query = urlencode(
        sorted(
            (key, value)
            for key, value in parse_qsl(parsed.query, keep_blank_values=True)
            if not key.lower().startswith("utm_") and key.lower() not in {"spm", "from", "source"}
        )
    )
    return parsed._replace(query=query, fragment="").geturl().rstrip("/")


def dedupe_rows(rows: Iterable[dict]):
    seen = set()
    for row in rows:
        key = canonical_url(row.get("url", "")) or clean_text(row.get("title", ""))
        if not key:
            yield row
            continue
        if key in seen:
            continue
        seen.add(key)
        yield row


def build_output_path(output_dir, topic):
    safe_topic = re.sub(r"[^\w\u4e00-\u9fff-]+", "_", topic).strip("_") or "crawl"
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
    return Path(output_dir) / f"{timestamp}_{uuid.uuid4().hex}_{safe_topic}.csv"


def build_log_path(log_dir, output_path):
    return Path(log_dir) / f"{output_path.stem}_task_log.json"


def write_task_log(log_path, payload):
    log_path.parent.mkdir(parents=True, exist_ok=True)
    with open(log_path, "w", encoding="utf-8") as file:
        json.dump(payload, file, ensure_ascii=False, indent=2)


def main():
    parser = argparse.ArgumentParser(description="按主题爬取白名单公开网页并输出 CSV。")
    parser.add_argument("--topic", required=True, help="主题关键词，多个关键词用空格分隔。")
    parser.add_argument("--whitelist", default=DEFAULT_WHITELIST, help="白名单 Excel 路径。")
    parser.add_argument("--output-dir", default=DEFAULT_OUTPUT_DIR, help="CSV 输出目录。")
    parser.add_argument("--log-dir", default=DEFAULT_LOG_DIR, help="任务日志输出目录。")
    parser.add_argument("--sites", default="", help="按 site_name 精确匹配多个站点，逗号分隔。")
    parser.add_argument("--limit-sites", type=int, default=0, help="只处理前 N 个启用站点，0 表示不限制。")
    parser.add_argument("--skip-sites", type=int, default=0, help="跳过前 N 个启用站点，适合分批测试。")
    parser.add_argument(
        "--run-mode",
        choices=["basic", "enhanced"],
        default="basic",
        help="运行路线：basic 为基础模式，不调用模型 API；enhanced 为增强模式，命中内容时弹窗调用模型 API。",
    )
    parser.add_argument(
        "--summary-mode",
        choices=["popup", "excerpt", "saved"],
        default=None,
        help="摘要模式覆盖：popup 表示首次摘要时弹窗填写 API；excerpt 表示直接截取正文前 200 字。",
    )
    parser.add_argument("--summary-provider", default="", help="增强模式一次性测试供应商预设，例如 DeepSeek V4 Flash。")
    parser.add_argument("--summary-api-url", default="", help="增强模式一次性测试完整 chat/completions 地址。")
    parser.add_argument("--summary-api-key", default="", help="增强模式一次性测试 API Key；不会写入文件。")
    parser.add_argument("--summary-model", default="", help="增强模式一次性测试模型名。")
    parser.add_argument(
        "--match-mode",
        choices=["any", "all"],
        default="any",
        help="关键词匹配模式：any 表示命中任一关键词即视为匹配（默认）；all 表示所有关键词必须全部命中。",
    )
    parser.add_argument("--matched-only", action="store_true", help="只输出关键词命中的结果。")
    parser.add_argument("--include-list-pages", action="store_true", help="将列表页本身也写入 CSV；默认只写详情页和失败状态。")
    parser.add_argument("--include-pdfs", action="store_true", help="允许抓取并提取 PDF 正文；需要安装 pypdf。")
    parser.add_argument("--date-from", default="", help="发布时间起始日期，格式 YYYY-MM-DD。")
    parser.add_argument("--date-to", default="", help="发布时间结束日期，格式 YYYY-MM-DD。")
    parser.add_argument("--no-dedupe", action="store_true", help="关闭输出 URL 去重。")
    parser.add_argument("--dry-run", action="store_true", help="只读取白名单和展示计划，不访问网络。")
    args = parser.parse_args()
    started_at = datetime.now()

    whitelist_path = Path(args.whitelist)
    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    sites = read_sites(whitelist_path)
    enabled_sites = [site for site in sites if site.enabled == "yes"]
    enabled_sites = filter_sites_by_names(enabled_sites, split_site_names(args.sites))
    if args.skip_sites > 0:
        enabled_sites = enabled_sites[args.skip_sites :]
    if args.limit_sites > 0:
        enabled_sites = enabled_sites[: args.limit_sites]
    keywords = split_keywords(args.topic)
    if not keywords:
        raise ValueError("主题不能为空。多个关键词请用空格分隔。")
    summary_mode = args.summary_mode or ("popup" if args.run_mode == "enhanced" else "excerpt")
    if args.summary_api_key or args.summary_provider or args.summary_api_url or args.summary_model:
        provider = PROVIDERS.get(args.summary_provider, {})
        api_url = args.summary_api_url or provider.get("api_url", "")
        model = args.summary_model or provider.get("model", "")
        if not api_url or not args.summary_api_key or not model:
            raise ValueError("使用摘要 API 参数时，需要 summary-provider/summary-api-url、summary-api-key 和 summary-model。")
        set_summary_config(api_url=api_url, api_key=args.summary_api_key, model=model)
        summary_mode = "api-args"
    else:
        set_summary_mode(summary_mode)
    date_from = parse_date(args.date_from, "--date-from")
    date_to = parse_date(args.date_to, "--date-to")

    print(f"主题: {args.topic}")
    print(f"关键词: {' / '.join(keywords)}")
    print(f"白名单: {whitelist_path}")
    print(f"启用网站数: {len(enabled_sites)}")
    print(f"运行路线: {args.run_mode}")
    print(f"摘要模式: {summary_mode}")
    if date_from or date_to:
        print(f"时间范围: {args.date_from or '不限'} 至 {args.date_to or '不限'}")

    if args.dry_run:
        for site in enabled_sites:
            print(f"- {site.site_name} | {site.url} | {site.domain} | {site.page_type} | max={site.max_detail_pages}")
        return

    output_path = build_output_path(output_dir, args.topic)
    print(f"匹配模式: {args.match_mode}")
    rows = iter_crawl_results(
        enabled_sites,
        args.topic,
        keywords,
        include_list_pages=args.include_list_pages,
        match_mode=args.match_mode,
        date_from=date_from,
        date_to=date_to,
        include_pdfs=args.include_pdfs,
    )
    if args.matched_only:
        rows = filter_matched(rows)
    if not args.no_dedupe:
        rows = dedupe_rows(rows)
    rows = list(rows)
    write_csv(rows, output_path)
    finished_at = datetime.now()
    log_path = build_log_path(args.log_dir, output_path)
    write_task_log(
        log_path,
        {
            "started_at": started_at.isoformat(timespec="seconds"),
            "finished_at": finished_at.isoformat(timespec="seconds"),
            "duration_seconds": round((finished_at - started_at).total_seconds(), 3),
            "topic": args.topic,
            "keywords": keywords,
            "whitelist": str(whitelist_path),
            "output_path": str(output_path),
            "enabled_site_count": len(enabled_sites),
            "matched_only": args.matched_only,
            "match_mode": args.match_mode,
            "date_from": args.date_from,
            "date_to": args.date_to,
            "include_pdfs": args.include_pdfs,
            "dedupe": not args.no_dedupe,
            "row_count": len(rows),
            "status_counts": {status: sum(1 for row in rows if row.get("status") == status) for status in sorted({row.get("status") for row in rows})},
            "site_counts": {site: sum(1 for row in rows if row.get("site_name") == site) for site in sorted({row.get("site_name") for row in rows})},
        },
    )
    print(f"输出: {output_path}")
    print(f"任务日志: {log_path}")


if __name__ == "__main__":
    main()
