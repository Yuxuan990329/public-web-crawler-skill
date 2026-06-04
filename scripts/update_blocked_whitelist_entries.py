from openpyxl import load_workbook


FILES = [
    "白名单网站模板.xlsx",
    "白名单网站模板_官方优化.xlsx",
    "白名单网站模板_栏目优化.xlsx",
]

BLOCKED_BY_ROBOTS = {"国家发改委", "中华人民共和国工业和信息化部"}


def update_file(path):
    workbook = load_workbook(path)
    sheet = workbook["whitelist"] if "whitelist" in workbook.sheetnames else workbook.active
    headers = [cell.value for cell in sheet[1]]
    indexes = {name: idx + 1 for idx, name in enumerate(headers)}

    rows_to_delete = []
    for row_num in range(2, sheet.max_row + 1):
        site_name = sheet.cell(row_num, indexes["site_name"]).value
        if site_name == "WGSN":
            rows_to_delete.append(row_num)
            continue
        if site_name in BLOCKED_BY_ROBOTS:
            sheet.cell(row_num, indexes["enabled"]).value = "no"
            sheet.cell(row_num, indexes["note"]).value = "robots.txt 不允许当前爬虫访问；保留来源记录，不进入自动抓取"

    for row_num in reversed(rows_to_delete):
        sheet.delete_rows(row_num, 1)

    workbook.save(path)
    print(f"{path}: removed_wgsn={len(rows_to_delete)}")


def main():
    for path in FILES:
        update_file(path)


if __name__ == "__main__":
    main()
