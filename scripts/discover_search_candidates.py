import argparse
import csv
import json
import re
import uuid
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from html import unescape
from urllib.parse import quote, urlencode, urljoin, urlparse
from urllib.request import Request, urlopen

from openpyxl import load_workbook

from crawl_whitelist import (
    DEFAULT_OUTPUT_DIR,
    DEFAULT_WHITELIST,
    USER_AGENT,
    can_fetch,
    fetch_html,
    is_same_or_subdomain,
    split_site_names,
    url_domain,
)
from extractors import clean_text, extract_page
from output_ownership import reserve_output_paths


SEARCH_TEMPLATES = {
    "yicai.com": ["https://www.yicai.com/search?keys={query}"],
    "chyxx.com": ["https://www.chyxx.com/search?word={query}"],
    "huaon.com": ["https://www.huaon.com/search?word={query}"],
    "qianzhan.com": [
        "https://bg.qianzhan.com/report/search/k-{query}.html",
        "https://t.qianzhan.com/search/?q={query}",
    ],
    "cbndata.com": ["https://www.cbndata.com/search?query={query}"],
    "36kr.com": ["https://36kr.com/search/articles/{query}"],
}

API_SEARCH_SITES = {"stats.gov.cn", "iresearch.com.cn", "cbndata.com", "djyanbao.com"}

AI_EXPANSION_KEYWORDS = [
    "AI",
    "人工智能",
    "大模型",
    "智能体",
    "AIGC",
    "生成式人工智能",
    "机器学习",
    "算力",
]

DIAGNOSTIC_QUERY_TERMS = {"回归测试", "修复后复测"}

TOPIC_SCOPE_MARKERS = ("行业", "产业", "市场", "发展", "变化", "趋势", "报告", "研究")

POSITIVE_PAGE_HINTS = [
    "news",
    "brief",
    "article",
    "industry",
    "research",
    "report",
    "channel",
    "information",
    "detail",
]

NEGATIVE_PAGE_HINTS = [
    "login",
    "register",
    "search",
    "rss",
    "english",
    "cart",
    "order",
    "about",
]


@dataclass
class Site:
    site_name: str
    url: str
    domain: str
    category: str
    enabled: str


def read_sites(path):
    workbook = load_workbook(path, data_only=True)
    sheet = workbook["whitelist"] if "whitelist" in workbook.sheetnames else workbook.active
    headers = [str(cell.value).strip() if cell.value is not None else "" for cell in sheet[1]]
    indexes = {name: idx for idx, name in enumerate(headers)}
    required = ["site_name", "url", "domain", "category", "enabled"]
    missing = [name for name in required if name not in indexes]
    if missing:
        raise ValueError(f"白名单缺少字段: {', '.join(missing)}")

    sites = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row or not row[indexes["url"]]:
            continue
        sites.append(
            Site(
                site_name=str(row[indexes["site_name"]] or "").strip(),
                url=str(row[indexes["url"]] or "").strip(),
                domain=str(row[indexes["domain"]] or "").strip(),
                category=str(row[indexes["category"]] or "").strip(),
                enabled=str(row[indexes["enabled"]] or "no").strip().lower(),
            )
        )
    return sites


def filter_sites_by_names(sites, names):
    if not names:
        return sites
    wanted = set(names)
    return [site for site in sites if site.site_name in wanted]


def split_keywords(value):
    return [part.strip() for part in re.split(r"[,，\s]+", value or "") if part.strip()]


def remove_diagnostic_terms(keywords):
    cleaned = []
    for keyword in keywords:
        for term in DIAGNOSTIC_QUERY_TERMS:
            keyword = keyword.replace(term, "").strip()
        if keyword:
            cleaned.append(keyword)
    return cleaned


def core_topic_term(keyword):
    """只返回可独立作为准入证据的实体词，范围/意图词不能单独得分。"""
    normalized = (keyword or "").strip()
    if not normalized:
        return ""
    marker_positions = [normalized.find(marker) for marker in TOPIC_SCOPE_MARKERS if marker in normalized]
    if marker_positions:
        prefix = normalized[: min(marker_positions)].strip()
        # “行业动态”“市场趋势”没有实体词，不能作为业务准入证据。
        return prefix if len(prefix) >= 2 else ""
    return normalized


def expand_topic_keywords(topic):
    keywords = []
    for keyword in remove_diagnostic_terms(split_keywords(topic)):
        core_term = core_topic_term(keyword)
        if not core_term:
            continue
        if core_term.lower() not in {item.lower() for item in keywords}:
            keywords.append(core_term)
    lowered = {keyword.lower() for keyword in keywords}
    if "ai" in lowered or "人工智能" in keywords:
        for keyword in AI_EXPANSION_KEYWORDS:
            if keyword.lower() not in lowered:
                keywords.append(keyword)
                lowered.add(keyword.lower())
    return keywords


def build_search_queries(topic, explicit_queries=""):
    if explicit_queries:
        queries = remove_diagnostic_terms(split_keywords(explicit_queries))
        if queries:
            return queries
    return expand_topic_keywords(topic)


def templates_for_site(site):
    domain = url_domain(site.url) or site.domain
    for registered_domain, templates in SEARCH_TEMPLATES.items():
        if domain == registered_domain or domain.endswith("." + registered_domain):
            return templates
    return []


def build_search_urls(site, queries):
    urls = []
    for template in templates_for_site(site):
        for query in queries:
            urls.append((query, template.format(query=quote(query))))
    return urls


def matched_keywords(text, keywords):
    haystack = (text or "").lower()
    return [keyword for keyword in keywords if keyword.lower() in haystack]


def publish_date_from_text(text):
    match = re.search(r"20\d{2}[-/.年]\d{1,2}[-/.月]\d{1,2}", text or "")
    return match.group(0) if match else ""


def normalize_date(value):
    match = re.search(r"(20\d{2})[-/.年]\s*(\d{1,2})[-/.月]\s*(\d{1,2})", value or "")
    if not match:
        return ""
    year, month, day = (int(part) for part in match.groups())
    return f"{year:04d}-{month:02d}-{day:02d}"


def strip_html(value):
    return clean_text(unescape(re.sub(r"<[^>]+>", " ", value or "")))


def link_context_snippet(html, href, fallback_text="", window=1200):
    if not html or not href:
        return fallback_text
    positions = []
    for needle in {href, href.replace("&", "&amp;")}:
        if not needle:
            continue
        index = html.find(needle)
        if index != -1:
            positions.append(index)
    if not positions:
        return fallback_text
    index = min(positions)
    start = max(0, index - window // 3)
    end = min(len(html), index + window)
    snippet = strip_html(html[start:end])
    return snippet or fallback_text


def score_candidate(url, title, keywords, snippet=""):
    title_matches = matched_keywords(title, keywords)
    url_lower = url.lower()
    parsed = urlparse(url)
    positives = [hint for hint in POSITIVE_PAGE_HINTS if hint in url_lower]
    negatives = [hint for hint in NEGATIVE_PAGE_HINTS if hint in url_lower]
    if parsed.path in {"", "/"}:
        negatives.append("root")
    score = len(title_matches) * 20 + len(positives) * 10 - len(negatives) * 30
    if re.search(r"/20\d{2}/|20\d{2}", url_lower):
        score += 5
    if len(clean_text(title)) >= 8:
        score += 3
    if not title_matches:
        score = min(score, 29)
    return score, title_matches, positives, negatives


def candidate_row(
    site,
    query,
    search_url,
    candidate_url="",
    title="",
    status="ok",
    error="",
    score=0,
    keywords=None,
    reason="",
    snippet="",
    publish_date="",
):
    keywords = keywords or []
    text = clean_text(title)
    clean_snippet = clean_text(snippet or title)
    return {
        "site_name": site.site_name,
        "category": site.category,
        "query": query,
        "search_url": search_url,
        "candidate_url": candidate_url,
        "title": text,
        "snippet": clean_snippet[:300],
        "publish_date": normalize_date(publish_date) or normalize_date(f"{text} {clean_snippet}"),
        "matched_keywords": " ".join(keywords),
        "preliminary_score": score,
        "status": status,
        "reason": reason,
        "error": error,
    }


def fetch_json(url, params=None, method="GET", referer=""):
    data = None
    full_url = url
    headers = {"User-Agent": USER_AGENT}
    if referer:
        headers["Referer"] = referer
    if params and method == "GET":
        full_url = f"{url}?{urlencode(params)}"
    elif params:
        data = urlencode(params).encode("utf-8")
        headers["Content-Type"] = "application/x-www-form-urlencoded; charset=UTF-8"
    request = Request(full_url, data=data, headers=headers, method=method)
    with urlopen(request, timeout=15) as response:
        return json.loads(response.read().decode("utf-8", errors="ignore"))


def discover_stats_api_search(site, query, keywords, limit):
    search_url = f"https://www.stats.gov.cn/search/s?qt={quote(query)}"
    if not can_fetch(search_url, USER_AGENT):
        return [candidate_row(site, query, search_url, status="skipped", error="robots.txt 不允许访问搜索页")]
    api_url = "https://api.so-gov.cn/query/s"
    data = fetch_json(
        api_url,
        {
            "siteCode": "bm36000002",
            "tab": "all",
            "qt": query,
            "page": "1",
            "pageSize": str(limit),
            "sort": "dateDesc",
        },
        method="POST",
        referer=search_url,
    )
    rows = []
    for item in data.get("resultDocs", [])[:limit]:
        detail = item.get("data") or {}
        candidate_url = detail.get("url") or ""
        if not candidate_url or not is_same_or_subdomain(candidate_url, site.domain):
            continue
        title = strip_html(detail.get("title") or detail.get("titleO") or "")
        snippet = strip_html(detail.get("summary") or "")
        score, hits, positives, negatives = score_candidate(candidate_url, title, keywords, snippet)
        if score <= 0:
            continue
        reason = f"api=so-gov; url_hints={','.join(positives)}"
        if negatives:
            reason += f"; negative={','.join(negatives)}"
        rows.append(
            candidate_row(
                site,
                query,
                search_url,
                candidate_url,
                title,
                score=score,
                keywords=hits,
                reason=reason,
                snippet=snippet,
                publish_date=detail.get("docDate") or "",
            )
        )
    return rows or [candidate_row(site, query, search_url, status="skipped", error="搜索 API 未发现可用候选链接")]


def discover_iresearch_api_search(site, query, keywords, limit):
    search_url = f"https://www.iresearch.com.cn/searchResults?keyword={quote(query)}"
    if not can_fetch(search_url, USER_AGENT):
        return [candidate_row(site, query, search_url, status="skipped", error="robots.txt 不允许访问搜索页")]
    data = fetch_json(
        "https://www.iresearch.com.cn/api/search/report/",
        {"keyword": query, "pageIndex": 1, "pageSize": limit, "order": 1},
        referer=search_url,
    )
    rows = []
    for item in (data.get("list") or data.get("List") or [])[:limit]:
        candidate_url = item.get("url") or ""
        if not candidate_url:
            report_id = item.get("id") or item.get("NewsId") or item.get("newsid")
            candidate_url = f"https://www.iresearch.com.cn/Detail/report?id={report_id}&isfree=0" if report_id else ""
        if not candidate_url:
            continue
        title = strip_html(item.get("hiliterTitle") or item.get("title") or item.get("Title") or "")
        snippet = strip_html(item.get("hkeyword") or item.get("shortcontent") or "")
        score, hits, positives, negatives = score_candidate(candidate_url, title, keywords, snippet)
        if score <= 0:
            continue
        reason = f"api=iresearch; url_hints={','.join(positives)}"
        if negatives:
            reason += f"; negative={','.join(negatives)}"
        rows.append(
            candidate_row(
                site,
                query,
                search_url,
                candidate_url,
                title,
                score=score,
                keywords=hits,
                reason=reason,
                snippet=snippet,
                publish_date=item.get("uptime") or "",
            )
        )
    return rows or [candidate_row(site, query, search_url, status="skipped", error="搜索 API 未发现可用候选链接")]


def discover_cbndata_api_search(site, query, keywords, limit):
    search_url = f"https://www.cbndata.com/search?query={quote(query)}"
    data = fetch_json(
        "https://api-next.cbndata.com/consumerstation/web/search/v1/contents",
        {
            "pageSize": str(limit),
            "pageIndex": "1",
            "contentType": "",
            "keyword": query,
        },
        referer=search_url,
    )
    rows = []
    payload = data.get("data") or {}
    for item in (payload.get("data") or [])[:limit]:
        item_id = item.get("id") or item.get("newId")
        item_type = item.get("type") or "information"
        if not item_id:
            continue
        if item_type not in {"information", "report", "activity"}:
            item_type = "information"
        candidate_url = f"https://www.cbndata.com/{item_type}/{item_id}"
        title = strip_html(item.get("title") or "")
        snippet = strip_html(" ".join([item.get("summary") or "", item.get("content") or ""]))
        score, hits, positives, negatives = score_candidate(candidate_url, title, keywords, snippet)
        if score <= 0:
            continue
        reason = f"api=cbndata; type={item_type}; url_hints={','.join(positives)}"
        if negatives:
            reason += f"; negative={','.join(negatives)}"
        rows.append(
            candidate_row(
                site,
                query,
                search_url,
                candidate_url,
                title,
                score=score,
                keywords=hits,
                reason=reason,
                snippet=snippet,
                publish_date=item.get("date") or "",
            )
        )
    return rows or [candidate_row(site, query, search_url, status="skipped", error="搜索 API 未发现可用候选链接")]


def discover_djyanbao_api_search(site, query, keywords, limit):
    search_url = f"https://www.djyanbao.com/report/search?q={quote(query)}"
    data = fetch_json(
        "https://api.djyanbao.com/api/report",
        {"q": query, "page": "1", "pageSize": str(limit)},
        referer=search_url,
    )
    rows = []
    payload = data.get("data") or {}
    for item in (payload.get("data") or [])[:limit]:
        item_id = item.get("id")
        if not item_id:
            continue
        candidate_url = f"https://www.djyanbao.com/report/detail?id={item_id}"
        title = strip_html(item.get("title") or "")
        highlight = item.get("highlight") or {}
        snippet_parts = []
        for value in highlight.values():
            if isinstance(value, list):
                snippet_parts.extend(str(part) for part in value)
            elif value:
                snippet_parts.append(str(value))
        snippet = strip_html(" ".join(snippet_parts))
        if not snippet:
            snippet = strip_html(" ".join([item.get("orgName") or "", item.get("authors") or ""]))
        score, hits, positives, negatives = score_candidate(candidate_url, title, keywords, snippet)
        if score <= 0:
            continue
        reason = f"api=djyanbao; public_preview_only; url_hints={','.join(positives)}"
        if negatives:
            reason += f"; negative={','.join(negatives)}"
        rows.append(
            candidate_row(
                site,
                query,
                search_url,
                candidate_url,
                title,
                score=score,
                keywords=hits,
                reason=reason,
                snippet=snippet,
                publish_date=item.get("publishAt") or "",
            )
        )
    return rows or [candidate_row(site, query, search_url, status="skipped", error="搜索 API 未发现可用候选链接")]


def discover_from_api_search(site, query, keywords, limit):
    domain = url_domain(site.url) or site.domain
    if domain == "stats.gov.cn" or domain.endswith(".stats.gov.cn"):
        return discover_stats_api_search(site, query, keywords, limit)
    if domain == "iresearch.com.cn" or domain.endswith(".iresearch.com.cn"):
        return discover_iresearch_api_search(site, query, keywords, limit)
    if domain == "cbndata.com" or domain.endswith(".cbndata.com"):
        return discover_cbndata_api_search(site, query, keywords, limit)
    if domain == "djyanbao.com" or domain.endswith(".djyanbao.com"):
        return discover_djyanbao_api_search(site, query, keywords, limit)
    return []


def discover_from_search(site, query, search_url, keywords, limit):
    if not is_same_or_subdomain(search_url, site.domain):
        return [candidate_row(site, query, search_url, status="skipped", error="搜索 URL 不匹配白名单域名")]
    if not can_fetch(search_url, USER_AGENT):
        return [candidate_row(site, query, search_url, status="skipped", error="robots.txt 不允许访问搜索页")]

    html = fetch_html(search_url)
    page = extract_page(search_url, html)
    rows = []
    seen = set()
    for href, link_text in page.links:
        absolute = urljoin(search_url, href)
        parsed = urlparse(absolute)
        if parsed.scheme not in {"http", "https"}:
            continue
        clean_url = parsed._replace(fragment="").geturl()
        if clean_url in seen:
            continue
        if not is_same_or_subdomain(clean_url, site.domain):
            continue
        if re.search(r"\.(pdf|doc|docx|xls|xlsx|zip|rar|jpg|jpeg|png|gif)(\?|$)", clean_url, re.I):
            continue
        snippet = link_context_snippet(html, href, link_text)
        score, hits, positives, negatives = score_candidate(clean_url, link_text, keywords, snippet)
        if score <= 0:
            continue
        seen.add(clean_url)
        reason = f"url_hints={','.join(positives)}"
        if negatives:
            reason += f"; negative={','.join(negatives)}"
        rows.append(
            candidate_row(
                site,
                query,
                search_url,
                clean_url,
                link_text or snippet[:80],
                score=score,
                keywords=hits,
                reason=reason,
                snippet=snippet,
            )
        )

    rows.sort(key=lambda row: int(row["preliminary_score"]), reverse=True)
    if not rows:
        return [candidate_row(site, query, search_url, status="skipped", error="搜索页未发现可用候选链接")]
    return rows[:limit]


def output_path(output_dir, topic):
    safe_topic = re.sub(r"[^\w\u4e00-\u9fff-]+", "_", topic).strip("_") or "search"
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
    return Path(output_dir) / f"{timestamp}_{uuid.uuid4().hex}_{safe_topic}_搜索候选.csv"


def main():
    parser = argparse.ArgumentParser(description="基于白名单站点的站内搜索页发现候选内容链接。")
    parser.add_argument("--topic", required=True, help="主题，例如 AI 或 人工智能。")
    parser.add_argument("--whitelist", default=DEFAULT_WHITELIST, help="白名单 Excel 路径。")
    parser.add_argument("--output-dir", default=DEFAULT_OUTPUT_DIR, help="CSV 输出目录。")
    parser.add_argument("--output", default="", help="由父任务指定的唯一 CSV 输出路径。")
    parser.add_argument("--sites", default="", help="按 site_name 精确匹配多个站点，逗号分隔。")
    parser.add_argument("--queries", default="", help="覆盖搜索词，逗号/空格分隔；默认由 topic 自动扩展。")
    parser.add_argument("--limit-sites", type=int, default=0, help="只处理前 N 个启用站点，0 表示不限制。")
    parser.add_argument("--limit-per-search", type=int, default=20, help="每个搜索 URL 最多输出 N 条候选。")
    parser.add_argument("--include-disabled", action="store_true", help="包含 enabled=no 的站点，仅用于诊断；仍尊重 robots.txt。")
    args = parser.parse_args()

    assigned_output = reserve_output_paths([args.output])[0] if args.output else None

    sites = read_sites(args.whitelist)
    if not args.include_disabled:
        sites = [site for site in sites if site.enabled == "yes"]
    sites = filter_sites_by_names(sites, split_site_names(args.sites))
    if args.limit_sites > 0:
        sites = sites[: args.limit_sites]

    queries = build_search_queries(args.topic, args.queries)
    keywords = expand_topic_keywords(args.topic)

    rows = []
    for site in sites:
        domain = url_domain(site.url) or site.domain
        if any(domain == api_domain or domain.endswith("." + api_domain) for api_domain in API_SEARCH_SITES):
            for query in queries:
                try:
                    rows.extend(discover_from_api_search(site, query, keywords, args.limit_per_search))
                except Exception as exc:
                    search_url = f"{site.url.rstrip('/')}/search"
                    rows.append(candidate_row(site, query, search_url, status="failed", error=str(exc)))
            continue
        search_urls = build_search_urls(site, queries)
        if not search_urls:
            rows.append(candidate_row(site, "", "", status="skipped", error="缺少该站点的搜索 URL 模板"))
            continue
        for query, search_url in search_urls:
            try:
                rows.extend(discover_from_search(site, query, search_url, keywords, args.limit_per_search))
            except Exception as exc:
                rows.append(candidate_row(site, query, search_url, status="failed", error=str(exc)))

    path = assigned_output if assigned_output else output_path(Path(args.output_dir), args.topic)
    path.parent.mkdir(parents=True, exist_ok=True)
    fields = [
        "site_name",
        "category",
        "query",
        "search_url",
        "candidate_url",
        "title",
        "snippet",
        "publish_date",
        "matched_keywords",
        "preliminary_score",
        "status",
        "reason",
        "error",
    ]
    with open(path, "w", newline="", encoding="utf-8-sig") as file:
        writer = csv.DictWriter(file, fieldnames=fields)
        writer.writeheader()
        for row in rows:
            writer.writerow(row)

    print(f"主题: {args.topic}")
    print(f"搜索词: {' / '.join(queries)}")
    print(f"处理站点数: {len(sites)}")
    print(f"候选行数: {len(rows)}")
    print(f"输出: {path}")


if __name__ == "__main__":
    main()
