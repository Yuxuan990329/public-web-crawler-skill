import argparse
import csv
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook


TARGET_SITES = ["中国家电网", "36氪", "洞见研报", "亿欧"]


def main():
    parser = argparse.ArgumentParser(description="将候选栏目链接写入新的白名单 Excel。")
    parser.add_argument("--whitelist", default="白名单网站模板.xlsx", help="原始白名单 Excel。")
    parser.add_argument("--candidates", required=True, help="候选栏目 CSV。")
    parser.add_argument("--sites", default="中国家电网,36氪,洞见研报,亿欧", help="要更新的 site_name，逗号分隔。")
    parser.add_argument("--top-n", type=int, default=3, help="每个站点最多写入前 N 条候选 URL。")
    parser.add_argument("--output", default="白名单网站模板_栏目优化.xlsx", help="输出白名单 Excel。")
    args = parser.parse_args()

    target_sites = [site.strip() for site in args.sites.replace("，", ",").split(",") if site.strip()]
    workbook = load_workbook(args.whitelist)
    sheet = workbook["whitelist"] if "whitelist" in workbook.sheetnames else workbook.active
    headers = [cell.value for cell in sheet[1]]
    indexes = {name: idx + 1 for idx, name in enumerate(headers)}

    candidates_by_site = defaultdict(list)
    with open(args.candidates, encoding="utf-8-sig") as file:
        for row in csv.DictReader(file):
            site_name = row.get("site_name", "")
            candidate_url = row.get("candidate_url", "")
            if site_name not in target_sites or not candidate_url:
                continue
            try:
                score = int(row.get("score", 0))
            except ValueError:
                score = 0
            candidates_by_site[site_name].append((score, row))

    for rows in candidates_by_site.values():
        rows.sort(key=lambda item: item[0], reverse=True)

    original_rows = []
    for values in sheet.iter_rows(min_row=2, values_only=True):
        if not values:
            continue
        original_rows.append(list(values))

    kept_rows = [row for row in original_rows if row[indexes["site_name"] - 1] not in target_sites]
    replacement_rows = []
    for site_name in target_sites:
        originals = [row for row in original_rows if row[indexes["site_name"] - 1] == site_name]
        if not originals:
            continue
        base = originals[0]
        selected = candidates_by_site.get(site_name, [])[: args.top_n]
        if not selected:
            row = list(base)
            row[indexes["enabled"] - 1] = "yes"
            note = row[indexes["note"] - 1] or ""
            row[indexes["note"] - 1] = f"{note}；未发现候选栏目，保留原入口，需浏览器人工核对"
            replacement_rows.append(row)
            continue

        for _score, candidate in selected:
            row = list(base)
            row[indexes["site_name"] - 1] = f"{site_name}-{candidate.get('link_text') or '候选栏目'}"
            row[indexes["url"] - 1] = candidate["candidate_url"]
            row[indexes["page_type"] - 1] = "list"
            row[indexes["enabled"] - 1] = "yes"
            row[indexes["note"] - 1] = f"候选栏目 score={candidate.get('score')}；来源：{site_name}"
            replacement_rows.append(row)

    sheet.delete_rows(2, sheet.max_row)
    for row in kept_rows + replacement_rows:
        sheet.append(row)

    workbook.save(args.output)
    print(args.output)


if __name__ == "__main__":
    main()
